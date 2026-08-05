"""
produce_paper_figures.py
========================
Generates all paper figures and tables. Formerly split across this file
and make_figures.py (an earlier draft with an ad-hoc plant-name-matching
colour fallback instead of a clean fuel-based lookup, and no colourblind
symbol overlay) - merged into this one file; make_figures.py is deleted.

    Step 0  : Workflow diagram (Figure 2) - 5-box monochrome flow schematic
    Step 1  : Merit order stacked bar - 2018-01-11, 2022-08-15, 2022-08-26, 2025-08-26
    Step 1b : Schram et al. (2019) validation (Figure 4) - MEI overlay for
              2018-01-11, plus schram_comparison_11jan2018.csv (Table 1's
              hourly companion). Moved here from model/plots.py - this is
              paper-specific validation content, not part of the
              distributable model.
    Step 2  : MEI time series for dispatch week (2022-08-26 to 2022-08-28)
    Step 3  : Dispatch detail: profit-DP vs Lexico-E
    Step 4  : Cumulative cashflow and CO2 - 2022 and 2025 (+ greedy-only and
              embodied-intensity-generality variants, Step 4b/4c)
    Step 5  : CO2 price over model horizon
    Step 6  : Annual KPI table (CSV + console)
    Step 7  : Console report
    Step 8  : Tables 1-4 (formatted .xlsx, for manuscript authoring)

Paper-specific analysis script (not part of the distributable model) - 
kept in support/, not model/. Imports the core layers from ../model/.

Run from storageemissions/:
    python support/produce_paper_figures.py
"""

import os, sys, warnings
warnings.filterwarnings("ignore")
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.lines as mlines
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "../model"))
from market import load_plants, compute_srmc, build_mei_curve, THERMAL_CO2, normalize_fuel
from run import FUEL_PRICE_MAP

BASE    = os.path.dirname(os.path.abspath(__file__))
BZNL    = os.path.join(BASE, "../data/raw/BZ_NL")
PROC    = os.path.join(BASE, "../data/processed/model")
FIG_DIR = os.path.join(PROC, "figures")
os.makedirs(FIG_DIR, exist_ok=True)

DISPATCH_WEEK_START = "2022-08-26"
DISPATCH_WEEK_END   = "2022-08-28"

# shared rcParams
plt.rcParams.update({
    "figure.dpi": 150, "savefig.dpi": 300, "savefig.bbox": "tight",
    "font.size": 9, "font.family": "sans-serif",
    "axes.grid": True, "grid.alpha": 0.22, "grid.linewidth": 0.6,
    "axes.spines.top": False, "axes.spines.right": False,
    "lines.linewidth": 1.3,
})

# colour scheme
FUEL_COLOR = {
    "coal":              "#6B3A2A",
    "biomass":           "#A8C66C",
    "gas_ccgt":          "#2166AC",
    "gas":               "#2166AC",
    "gas_ccgt_chp":      "#74ADD1",
    "gas_ocgt":          "#E65100",
    "gas_engine_chp":    "#A6CEE3",
    "blast_furnace_gas": "#762A83",
    "nuclear":           "#F6B50A",
    "res":               "#4DAF4A",
    "wind":              "#4DAF4A",
    "solar":             "#4DAF4A",
    "other":             "#AAAAAA",
}
FUEL_SYMBOL = {
    "coal":              "x",    # cross
    "biomass":           "s",    # square
    "gas_ccgt":          "o",    # circle
    "gas":               "o",    # circle
    "gas_ccgt_chp":      "^",    # triangle up
    "gas_ocgt":          "D",    # diamond
    "gas_engine_chp":    "v",    # triangle down
    "blast_furnace_gas": "*",    # star
    "nuclear":           None,     # zero emissions - skip
    "res":               None,
    "wind":              None,
    "solar":             None,
    "other":             ".",
}
COMP_STYLE  = {"fuel": (None, "white", 0.92), "co2": ("///", "#333", 0.85), "varom": ("...", "#333", 0.88)}
COMP_LABELS = {"fuel": "Fuel cost (p/η)", "co2": "CO2/ETS cost", "varom": "Var. O&M"}
DEMO_GREY   = "#888888"

_DISPLAY_MAP = {
    "Coal": "coal", "Biomass": "biomass", "Gas CCGT": "gas_ccgt",
    "Gas CCGT CHP": "gas_ccgt_chp", "Gas OCGT": "gas_ocgt",
    "Gas Engine CHP": "gas_engine_chp", "Blast furnace gas": "blast_furnace_gas",
    "Nuclear": "nuclear", "Solar + Wind": "res",
}
_CANONICAL = {"gas":"gas_ccgt","coal":"coal","biomass":"biomass",
              "blast_furnace_gas":"blast_furnace_gas","nuclear":"nuclear",
              "wind":"res","solar":"res"}

def display_fuel(raw):
    s = str(raw).strip()
    if s in _DISPLAY_MAP: return _DISPLAY_MAP[s]
    c = normalize_fuel(s)
    return _CANONICAL.get(c, "other")

# Plant name abbreviations for bar labels
ABBREV = {
    "Eemshaven A/B (RWE)":         "Eems A/B",
    "Maasvlakte 3 (MPP3)":         "MPP3",
    "Rotterdam 1 (Onyx)":          "RTD 1 ONYX",
    "Amer (Amercentrale)":         "Amer 9",
    "Hemweg 8 (coal, ret. 2020)":  "Hem 8",
    "Flevo 4/5 (Maxima)":          "Flevo",
    "Diemen 34 CCGT":              "Diem34",
    "Enecogen 10/20":              "Enecogen",
    "Hemweg 9":                    "Hem 9",
    "Sloecentrale 10/20":          "Sloe",
    "Claus C 2/3/4/5":             "Claus",
    "Eemshaven 10/20/30 (Magnum)": "Magnum",
    "Moerdijk 2 CCGT":             "Mrd2",
    "Maasstroom Rijnmond 2":       "Rnmnd2",
    "Eems CCGT 4/5/6/7":           "Eems 4/5/6/7",
    "Elsta Cogeneration":          "Elsta",
    "Moerdijk 1 CHP":              "Mrd 1",
    "Rijnmond 1":                  "Rnmnd1",
    "Diemen 33 CHP":               "Diemen",
    "Delesto 2":                   "Delesto",
    "Lage Weide 6":                "Lage Weide 6",
    "Merwedekanaal Pegus 12":      "Merwede",
    "NAM Schoonebeek":             "NAM",
    "Den Haag Power station":      "DenH",
    "RoCa":                        "RoCa",
    "Pergen 1/2":                  "Perg",
    "Swentibold Chemelot":         "Swntbld",
    "Velsen 25":                   "Vls25",
    "Velsen 24":                   "Vls24",
    "Bergum 10GT/20GT Friesland":  "Bergum",
    "Eems 20":                     "Eems20",
    "Borssele 30":                 "Bors",
    "RES (wind + solar)":          "RES",
}

STRAT_COLOR = {
    # Validated against scripts/validate_palette.js (dataviz skill), AND
    # chosen for luminance contrast within each pair that visually overlays
    # (profit_dp exactly overlays lexico_profit_dp; emission_dp exactly
    # overlays lexico_emissions_dp in most years) - hue distance alone was
    # not enough there: the original green/emission_dp vs blue/lexico
    # pairing had matching hue-CVD-separation but almost no luminance gap
    # (relative luminance 0.162 vs 0.188), so a thin dashed line on top of
    # a thin solid line of similar lightness read as one blended line even
    # though the colors are theoretically well separated. Each pair below
    # now also differs materially in luminance:
    #   profit_dp/lexico_profit_dp:      0.216 vs 0.073 (gap 0.143)
    #   emission_dp/lexico_emissions_dp: 0.188 vs 0.278 (gap 0.090)
    # profit_greedy stays neutral grey (deliberately achromatic; it also
    # carries its own line style + legend label, so it doesn't need to
    # compete as a categorical hue).
    "profit_dp":           "#E34948",   # red
    "emission_dp":         "#2A78D6",   # blue
    "lexico_emissions_dp": "#EB6834",   # orange
    "lexico_profit_dp":    "#4A3AA7",   # violet
    "profit_greedy":       "#757575",   # grey
}
STRAT_LABEL = {
    "profit_dp":           "Profit max (DP)",
    "emission_dp":         "Emission min (DP)",
    "lexico_emissions_dp": "Lexico-E: emission primary",
    "lexico_profit_dp":    "Lexico-P: profit primary",
    "profit_greedy":       "Greedy benchmark",
}
STRAT_LS = {"profit_dp":"-","emission_dp":"-","lexico_emissions_dp":"--",
            "lexico_profit_dp":"--","profit_greedy":"-"}

saved_figures = []
skipped_figures = []
failed_figures = []


def savefig(fig, stem):
    for ext in ["png", "pdf"]:
        p = os.path.join(FIG_DIR, f"{stem}.{ext}")
        try:
            fig.savefig(p)
            if ext == "png":
                saved_figures.append(p)
        except PermissionError:
            failed_figures.append((p, "PermissionError - close file in viewer"))
    plt.close(fig)


# helpers

def load_prices():
    raw = pd.read_csv(os.path.join(BZNL, "commodity_prices_BZNL.csv"),
                      parse_dates=["Datum"], index_col="Datum")
    raw = raw.rename(columns={"Kolen": "Coal", "blast furnace gas": "Blast_furnace_gas"})
    idx = pd.date_range(raw.index.min(), raw.index.max() + pd.Timedelta(hours=23), freq="h")
    return raw.reindex(idx).ffill()


def load_dam():
    raw = pd.read_csv(os.path.join(BZNL, "BZ_NL.csv"))
    raw["dt"] = (raw["Date"].str.split(" - ").str[0]
                 .str.replace(r"\s*\(.*?\)", "", regex=True).str.strip())
    raw["dt"] = pd.to_datetime(raw["dt"], format="%d/%m/%Y %H:%M:%S")
    return raw.set_index("dt").sort_index()["Price"].astype(float).resample("h").mean()


def load_amer_mix():
    return pd.read_csv(os.path.join(BZNL, "amer_fuel_mix.csv"),
                       comment="#", parse_dates=["date_from"])


def apply_amer_blend(plants, ts, amer_mix):
    from market import apply_cofiring_blend
    return apply_cofiring_blend(plants, "Amer (Amercentrale)", ts, amer_mix)


def build_merit_order_df(date_str):
    """Return sorted merit-order DataFrame for a given date string YYYY-MM-DD."""
    year = int(date_str[:4])
    ts   = pd.Timestamp(date_str)

    stack = os.path.join(BZNL, f"meritorder_NL_{year}.csv")
    if not os.path.exists(stack):
        stack = os.path.join(BZNL, "meritorder_NL.csv")

    raw_fuels = (pd.read_csv(stack).set_index("name")["fuel"].to_dict())
    plants    = load_plants(stack)
    amer_mix  = load_amer_mix()
    plants    = apply_amer_blend(plants, ts, amer_mix)

    prices    = load_prices()
    price_row = prices.loc[ts.floor("h")]

    prices_df = pd.DataFrame([price_row], index=[ts])
    srmc      = compute_srmc(plants, prices_df, FUEL_PRICE_MAP)
    srmc_ts   = srmc.loc[ts]

    # Tiebreaker: RES before nuclear
    sk = srmc_ts.copy()
    for p in sk.index:
        if normalize_fuel(str(plants.loc[p, "fuel"]) if p in plants.index else "") in ("wind","solar"):
            sk[p] -= 1e-6
    srmc_ts = srmc_ts.loc[sk.sort_values().index]

    rows = []
    for name, sv in srmc_ts.items():
        p = plants.loc[name]
        cap = float(p.get("capacity_mw", 0))
        if cap == 0: continue  # RES included: SRMC=0 shows as hatched green bar
        eta = float(p["efficiency"])
        fuel_raw = raw_fuels.get(name, str(p["fuel"]))
        canon = normalize_fuel(str(p["fuel"]))
        pc = FUEL_PRICE_MAP.get(canon)
        fc = float(price_row[pc]) / eta if pc and pc in price_row.index else 0.0
        # CO2 intensity
        co2_col = p.get("co2_kg_per_kwh", float("nan"))
        try:
            eps = float(co2_col) if not pd.isna(co2_col) else THERMAL_CO2.get(canon,0.0)/eta
        except Exception:
            eps = THERMAL_CO2.get(canon, 0.0) / eta
        if "Amer" in name:
            mr = amer_mix[amer_mix["date_from"] <= ts].sort_values("date_from")
            if not mr.empty:
                cs = float(mr.iloc[-1]["coal_share"])
                eps = cs * THERMAL_CO2["coal"] / eta
        cc = float(price_row.get("CO2", 0.0)) * eps
        vom = float(p.get("var_om_eur_mwh", 0.0))
        rows.append({
            "name": name, "fuel_raw": fuel_raw, "cap_mw": cap,
            "srmc": round(sv, 3), "fuel_cost": round(fc,3),
            "co2_cost": round(cc,3), "var_om": round(vom,3),
            "em_kgmwh": round(eps*1000,0),
        })

    return pd.DataFrame(rows), price_row


# Step 1: merit order stacked bar figures

def merit_order_figure(date_str, subtitle, fname, emissions_overlay=True):
    print(f"\n  Building merit order for {date_str} ...")
    df, price_row = build_merit_order_df(date_str)

    co2_p  = float(price_row.get("CO2",  float("nan")))
    gas_p  = float(price_row.get("Gas",  float("nan")))
    coal_p = float(price_row.get("Coal", float("nan")))
    print(f"    CO2={co2_p:.2f} EUR/t | Gas={gas_p:.2f} EUR/MWh_th | Coal={coal_p:.2f} EUR/MWh_th")
    subtitle = subtitle or (
        f"Market prices: CO2 = {co2_p:.2f} EUR/t | "
        f"Gas = {gas_p:.2f} EUR/MWh_th | Coal = {coal_p:.2f} EUR/MWh_th"
    )

    max_srmc  = df["srmc"].max()
    total_cap = df["cap_mw"].sum()
    zero_bar  = max_srmc * 0.025

    fig, ax = plt.subplots(figsize=(13, 6))
    ax.spines["right"].set_visible(False)
    cumcap = 0.0
    bar_centers = []
    plant_nums = {}
    bar_number = 1

    for _, row in df.iterrows():
        cap  = row["cap_mw"]
        srmc = row["srmc"]
        disp = display_fuel(row["fuel_raw"])
        clr  = FUEL_COLOR.get(disp, FUEL_COLOR["other"])
        xc   = cumcap + cap / 2
        is_z = srmc < 0.01
        bh   = zero_bar if is_z else srmc

        if is_z:
            ax.bar(xc, bh, width=cap, bottom=0, color=clr,
                   edgecolor="white", lw=0.3, align="center", hatch="///", alpha=0.85)
        else:
            # Stacked: fuel | co2 | varom
            bot = 0.0
            for seg, val in [("fuel", row["fuel_cost"]),
                              ("co2",  row["co2_cost"]),
                              ("varom",row["var_om"])]:
                if val <= 0: continue
                h, ec, al = COMP_STYLE[seg]
                ax.bar(xc, val, width=cap, bottom=bot, color=clr,
                       edgecolor=ec, lw=0.3, align="center", hatch=h, alpha=al)
                bot += val
            bh = bot

        # Schram-style numbering instead of plant names
        ax.text(xc, bh + max_srmc * 0.008, str(bar_number),
            ha="center", va="bottom", fontsize=5, color="black", clip_on=True)
        plant_nums[bar_number] = str(row["name"])
        bar_number += 1
        bar_centers.append((xc, disp, float(row["em_kgmwh"])))
        cumcap += cap

    ax.set_xlim(0, total_cap * 1.01)
    ax.set_ylim(0, max_srmc * 1.30)
    ax.spines["left"].set_position(("data", 0))
    ax.set_xlabel("Cumulative installed capacity [MW]")
    ax.set_ylabel("SRMC [EUR/MWh_e]")
    ax.set_title(f"Merit order – {pd.Timestamp(date_str).strftime('%d-%m-%Y')}\n{subtitle}",
                 fontsize=9)

    sym_handles = []
    if emissions_overlay:
        ax2 = ax.twinx()
        ax2.spines["right"].set_visible(True)
        ax2.spines["top"].set_visible(False)
        max_em = max((em for _, _, em in bar_centers if em > 0), default=1000)
        ax2.set_ylim(0, max_em * 1.25)
        ax2.set_ylabel("CO2 intensity [kg CO2/MWh_e]", color="#444444")
        ax2.tick_params(axis="y", colors="#444444")

        seen_sym = set()
        for x_ctr, fuel, em in bar_centers:
            if em < 1.0:
                continue
            sym = FUEL_SYMBOL.get(fuel)
            if sym is None:
                continue
            ax2.plot([x_ctr, x_ctr], [0, em], linestyle=":", color="#555555",
                     lw=1.0, alpha=0.6, zorder=5)
            ax2.plot(x_ctr, em, marker=sym, color="#333333", markersize=5,
                     markeredgewidth=0.8, zorder=6, linestyle="none")
            if fuel not in seen_sym:
                lbl = {
                    "coal": "Coal",
                    "biomass": "Biomass",
                    "gas_ccgt": "Gas CCGT",
                    "gas_ccgt_chp": "Gas CCGT CHP",
                    "gas_ocgt": "Gas OCGT",
                    "gas_engine_chp": "Gas Engine CHP",
                    "blast_furnace_gas": "Blast furnace gas",
                    "other": "Other",
                }.get(fuel, fuel.replace("_", " ").title())
                sym_handles.append(
                    mlines.Line2D([0], [0], linestyle="none", marker=sym,
                                  color="#333333", markersize=5,
                                  markeredgewidth=0.8, label=lbl)
                )
                seen_sym.add(fuel)

    # Legend: fuel colours
    fuel_legend = [
        ("coal",              "Coal"),
        ("biomass",           "Biomass"),
        ("gas_ccgt",          "Gas CCGT"),
        ("gas_ccgt_chp",      "Gas CCGT CHP"),
        ("gas_ocgt",          "Gas OCGT"),
        ("blast_furnace_gas", "Blast furnace gas"),
        ("nuclear",           "Nuclear (SRMC=0)"),
        ("res",               "RES (SRMC=0)"),
    ]
    # Only show fuels that appear in the stack
    fuels_present = {display_fuel(r["fuel_raw"]) for _, r in df.iterrows()}
    handles: list = [mpatches.Patch(color=FUEL_COLOR[f], label=l)
                     for f, l in fuel_legend if f in fuels_present]
    if sym_handles:
        handles.append(mlines.Line2D([], [], linestyle="none", label=" "))
        handles += sym_handles
    # Component legend
    handles.append(mlines.Line2D([], [], linestyle="none", label=" "))
    for seg, lbl in COMP_LABELS.items():
        h, ec, al = COMP_STYLE[seg]
        handles.append(mpatches.Patch(
            facecolor=DEMO_GREY,
            edgecolor=ec if ec != "white" else "#555",
            hatch=h, alpha=al, label=lbl))

    ax.legend(handles=handles, loc="upper left", fontsize=7,
              framealpha=0.9, bbox_to_anchor=(0.0, 1.0))
    fig.tight_layout()
    savefig(fig, fname)

    # Save number-to-plant mapping, matching Schram validation style.
    leg = pd.DataFrame(list(plant_nums.items()), columns=["number", "plant_name"])
    leg.to_csv(os.path.join(PROC, f"{fname}_legend.csv"), index=False)

    return co2_p, gas_p, coal_p


# Step 1b: Schram et al. (2019) validation (Figure 4, Table 1's hourly companion)
# Reference: Schram, W., Lampropoulos, I., AlSkaif, T. & Van Sark, W. (2019).
# On the use of average versus marginal emission factors. SMARTGREENS 2019,
# pp. 187-193. https://doi.org/10.5220/0007765701870193

SCHRAM_VALIDATION_DATE = "2018-01-11"

# Schram et al. (2019), Figure 2 - digitised MEI curve [kg CO2/MWh_e], hour 0-24.
SCHRAM_DIGITISED = [
    920, 920, 920, 895, 895, 920,   #  0- 5: overnight coal
    330, 400, 450, 480, 490, 480,   #  6-11: gas enters, rises to peak
    450, 450, 450, 450, 450, 450,   # 12-17: afternoon gas plateau ~450
    450, 450, 400, 350, 350,        # 18-22: step down
    920, 920,                       # 23-24: coal returns
]
assert len(SCHRAM_DIGITISED) == 25


def schram_validation_figure():
    """Figure 4: this model's MEI curve for 11 Jan 2018 vs Schram et al.
    (2019)'s digitised curve. Also writes schram_comparison_11jan2018.csv,
    the hour-by-hour companion to Table 1."""
    date = SCHRAM_VALIDATION_DATE
    year = int(date[:4])
    day  = pd.date_range(f"{date} 00:00", f"{date} 23:00", freq="h")

    stack = os.path.join(BZNL, f"meritorder_NL_{year}.csv")
    raw_fuels = pd.read_csv(stack).set_index("name")["fuel"].to_dict()
    plants    = load_plants(stack)
    amer_mix  = load_amer_mix()
    plants    = apply_amer_blend(plants, pd.Timestamp(date), amer_mix)

    prices  = load_prices().loc[day]
    dam_day = load_dam().reindex(day)

    srmc, marginal, mei = build_mei_curve(plants, prices, dam_day, FUEL_PRICE_MAP)
    mei_kg = mei * 1000.0

    hour_fuels = [display_fuel(raw_fuels.get(str(marginal.get(ts, "")), "")) for ts in day]

    fig, ax1 = plt.subplots(figsize=(7, 4))
    ax2 = ax1.twinx()
    ax2.spines["right"].set_visible(True)

    for h in range(24):
        ax1.plot([h, h + 1], [mei_kg.iloc[h], mei_kg.iloc[h]],
                 color=FUEL_COLOR.get(hour_fuels[h], FUEL_COLOR["other"]),
                 lw=2.5, solid_capstyle="butt")

    ax1.step(range(25), SCHRAM_DIGITISED, where="post", color="#C62828", lw=1.5, ls=":")
    ax2.step(range(24), list(dam_day.values), where="post", color="#888888", lw=1.0, ls="--", alpha=0.7)

    ax1.set_xlabel("Hour of day")
    ax1.set_ylabel("MEI [kg CO2/MWh_e]")
    ax2.set_ylabel("Day-ahead price [EUR/MWh]", color="#888888")
    ax1.set_xlim(0, 24); ax1.set_xticks(range(0, 25, 4))
    ax1.set_ylim(0, 1100)

    fuel_handles = [
        mlines.Line2D([0], [0], color=FUEL_COLOR.get(f, FUEL_COLOR["other"]), lw=2.5,
                      label=f.replace("_", " ").title())
        for f in sorted(set(hour_fuels))
    ]
    line_handles = [
        mlines.Line2D([0], [0], color="#C62828", ls=":", lw=1.5, label="Schram et al. (2019)"),
        mlines.Line2D([0], [0], color="#888888", ls="--", lw=1.0, alpha=0.7, label="DAM price"),
    ]
    # Legend placed outside the axes (left margin) rather than inside -
    # MEI values span the full y-range including the overnight coal level,
    # so any inside corner risks the legend box occluding a data segment.
    ax1.legend(handles=fuel_handles + line_handles, fontsize=7, framealpha=0.9,
               loc="center right", bbox_to_anchor=(-0.18, 0.5),
               bbox_transform=ax1.transAxes, borderaxespad=0)
    ax1.set_title("Schram Validation, 11-1-2018\nMEI: this model vs Schram et al. (2019)")
    fig.subplots_adjust(left=0.35)

    savefig(fig, "fig_schram_validation")
    print("    Saved: fig_schram_validation.pdf / .png")

    schram_marginal = {h: "coal" if SCHRAM_DIGITISED[h] > 600 else "gas" for h in range(24)}
    rows = []
    for h in range(24):
        our     = str(marginal.get(day[h], ""))
        our_mei = float(mei_kg.iloc[h])
        sch_mei = SCHRAM_DIGITISED[h]
        diff    = our_mei - sch_mei
        if abs(diff) < 30:                                  reason = "Agreement"
        elif "Velsen" in our:                                reason = "BFG emission factor"
        elif "Amer" in our:                                  reason = "Amer biomass cofiring"
        elif hour_fuels[h] == "coal" and sch_mei < 600:      reason = "Fuel price vintage"
        else:                                                reason = "Merit order position"
        rows.append({
            "hour": h,
            "dam_price": round(float(dam_day.iloc[h]), 2),
            "schram_marginal": schram_marginal[h],
            "schram_mei_kgmwh": sch_mei,
            "our_marginal": our,
            "our_mei_kgmwh": round(our_mei, 1),
            "difference_kg": round(diff, 1),
            "reason": reason,
        })
    csv_path = os.path.join(PROC, "schram_comparison_11jan2018.csv")
    pd.DataFrame(rows).to_csv(csv_path, index=False)
    print("    Saved: schram_comparison_11jan2018.csv")


# Step 2: MEI time series for dispatch week

def mei_dispatch_week_figure():
    print("\n  Building MEI time series for dispatch week ...")
    # Load from saved annual results (has `mei` column in t/MWh_e = kg/kWh_e)
    disp_csv = os.path.join(PROC, "annual_results_2022_profit_dp.csv")
    disp = pd.read_csv(disp_csv, index_col=0, parse_dates=True)
    mei  = disp["mei"].loc[DISPATCH_WEEK_START:DISPATCH_WEEK_END] * 1000.0  # kg/MWhe

    # Marginal fuel from mei_curve.csv if available
    mc_path = os.path.join(PROC, "mei_curve.csv")
    if os.path.exists(mc_path):
        mc   = pd.read_csv(mc_path, index_col=0, parse_dates=True)
        mc   = mc.loc[DISPATCH_WEEK_START:DISPATCH_WEEK_END]
        marg = mc["marginal_plant"] if "marginal_plant" in mc.columns else None
    else:
        marg = None

    # Colour each hour by marginal fuel
    # Fallback: derive from MEI level
    def mei_to_color(m_val, plant_name=""):
        if plant_name:
            # Try to infer from plant name
            n = str(plant_name).lower()
            if "velsen" in n: return FUEL_COLOR["blast_furnace_gas"]
            if any(x in n for x in ["ams","borssele","nuclear"]): return FUEL_COLOR["nuclear"]
            if any(x in n for x in ["res","wind","solar"]): return FUEL_COLOR["res"]
            if any(x in n for x in ["coal","eemshaven a","maasvlak","rotterdam","hemweg 8","amer"]): return FUEL_COLOR["coal"]
            if any(x in n for x in ["ccgt chp","diemen 33","lage weide","merwedekanaal","roca","elsta","delesto","rijnmond","pergen","swentibold","moerdijk 1","nam"]): return FUEL_COLOR["gas_ccgt_chp"]
            if any(x in n for x in ["bergum","eems 20","ocgt"]): return FUEL_COLOR["gas_ocgt"]
        # Fallback by MEI level
        if m_val < 5:   return FUEL_COLOR["res"]
        if m_val < 50:  return FUEL_COLOR["nuclear"]
        if m_val < 100: return FUEL_COLOR["blast_furnace_gas"]
        if m_val < 500: return FUEL_COLOR["gas_ccgt"]
        return FUEL_COLOR["coal"]

    dam = load_dam().loc[DISPATCH_WEEK_START:DISPATCH_WEEK_END]

    fig, ax1 = plt.subplots(figsize=(12, 4))
    ax2 = ax1.twinx()
    ax1.set_zorder(1)
    ax2.set_zorder(2)
    ax2.patch.set_visible(False)
    ax2.spines["right"].set_visible(True)
    ax2.spines["top"].set_visible(False)

    max_mei = float(mei.max()) if len(mei) else 0.0
    zero_bar = max(max_mei * 0.025, 1.0)

    for i in range(len(mei) - 1):
        ts_i = mei.index[i]
        plant = str(marg.iloc[i]) if marg is not None and i < len(marg) else ""
        m_val = float(mei.iloc[i])
        vis_val = zero_bar if m_val < 0.01 else m_val
        clr   = mei_to_color(m_val, plant)
        ax1.fill_between([ts_i, mei.index[i+1]],
                         [0, 0], [vis_val, vis_val],
                         color=clr, alpha=0.65, step="post", zorder=2)
    ax1.step(mei.index, mei.values, where="post", color="black", lw=0.8, alpha=0.5)
    ax2.step(dam.index, dam.values, where="post",
             color="#1F77B4", lw=1.2, ls="--", alpha=0.95, label="DAM price", zorder=10)

    # Day boundary lines
    for d in pd.date_range(DISPATCH_WEEK_START, DISPATCH_WEEK_END, freq="D"):
        ax1.axvline(d, color="black", lw=0.4, ls=":", alpha=0.5)

    ax1.set_ylabel("MEI [kg CO₂/MWh_e]")
    ax2.set_ylabel("DAM price [EUR/MWh]", color="#1F77B4")
    ax2.tick_params(axis="y", colors="#1F77B4")
    ax1.set_title(f"Marginal Emission Intensity – {DISPATCH_WEEK_START} to {DISPATCH_WEEK_END}, Dutch bidding zone")
    ax1.set_ylim(0)

    legend_f = [("coal","Coal"), ("gas_ccgt","Gas CCGT"),
                ("gas_ccgt_chp","Gas CCGT CHP"), ("blast_furnace_gas","BFG"),
                ("nuclear","Nuclear"), ("res","RES")]
    patches  = [mpatches.Patch(color=FUEL_COLOR[f], alpha=0.65, label=l) for f,l in legend_f]
    patches += [mlines.Line2D([],[],color="#1F77B4",ls="--",lw=1.2,label="DAM price")]
    ax1.legend(handles=patches, loc="upper left", fontsize=7, framealpha=0.9)
    fig.tight_layout()
    savefig(fig, "fig_mei_dispatch_week")


# Step 3: dispatch detail figure

def dispatch_detail_figure():
    print("\n  Building dispatch detail figure ...")
    from run import UNIT

    dam  = load_dam().loc[DISPATCH_WEEK_START:DISPATCH_WEEK_END]
    strats = [("profit_dp", "Profit max (DP)"),
              ("lexico_emissions_dp", "Lexico-E: emission primary")]

    fig, axes = plt.subplots(len(strats), 1, figsize=(12, 9), sharex=True)
    fig.subplots_adjust(right=0.78)

    for ax, (s, title) in zip(axes, strats):
        csv_path = os.path.join(PROC, f"annual_results_2022_{s}.csv")
        disp = pd.read_csv(csv_path, index_col=0, parse_dates=True)
        disp = disp.loc[DISPATCH_WEEK_START:DISPATCH_WEEK_END]
        mei  = disp["mei"] * 1000.0   # kg/MWhe

        # Left axis: bars + SOC (normalised -1 to +1)
        bw = 1/24
        ax.bar(disp.index,  disp["charge_mw"]    / UNIT.p_max_mw,
               color="#2166AC", alpha=0.75, width=bw, label="Charge [p.u.]")
        ax.bar(disp.index, -disp["discharge_mw"] / UNIT.p_max_mw,
               color="#D32F2F", alpha=0.75, width=bw, label="Discharge [p.u.]")
        soc_norm = disp["soc_mwh"] / UNIT.e_cap_mwh
        ax.plot(disp.index, soc_norm * 2 - 1, color="black", lw=1.2, ls=":", label="SOC [0–1→axis]")
        ax.axhline(0, color="black", lw=0.4)
        ax.set_ylim(-1.15, 1.15)
        ax.set_yticks([-1, -0.5, 0, 0.5, 1])
        ax.set_ylabel("Charge/Discharge [-]  |  SOC", fontsize=8)
        ax.set_title(title)

        # Day boundaries
        for d in pd.date_range(DISPATCH_WEEK_START, DISPATCH_WEEK_END, freq="D"):
            ax.axvline(d, color="black", lw=0.4, ls=":", alpha=0.4)

        # Right inner: DAM price
        ax_dam = ax.twinx()
        ax_dam.spines["right"].set_visible(True); ax_dam.spines["top"].set_visible(False)
        l1, = ax_dam.plot(dam.index, dam.values, color="#888", lw=1.2, ls="-", alpha=0.8,
                          label="DAM price [EUR/MWh]")
        ax_dam.set_ylabel("DAM price [EUR/MWh]", color="#555", fontsize=8)
        ax_dam.tick_params(axis="y", colors="#555", labelsize=7)

        # Right outer: MEI
        ax_mei = ax.twinx()
        ax_mei.spines["right"].set_position(("outward", 65))
        ax_mei.spines["right"].set_visible(True); ax_mei.spines["top"].set_visible(False)
        l2, = ax_mei.plot(mei.index, mei.values, color="#00838F", lw=1.0, ls="--", alpha=0.85,
                          label="MEI [kg CO2/MWh_e]")
        ax_mei.set_ylabel("MEI [kg CO2/MWh_e]", color="#00838F", fontsize=8)
        ax_mei.tick_params(axis="y", colors="#00838F", labelsize=7)

        h_l, l_l = ax.get_legend_handles_labels()
        ax.legend(h_l + [l1, l2], l_l + [l1.get_label(), l2.get_label()],
                  loc="upper left", fontsize=7, framealpha=0.9)

    axes[-1].set_xlabel(f"Date (2022)")
    fig.suptitle(f"Dispatch detail: {DISPATCH_WEEK_START} to {DISPATCH_WEEK_END}  (gas crisis: coal and gas charging)")
    savefig(fig, "fig_dispatch_detail_week")


# Step 4: cumulative cashflow and CO2 figures

STRATS_ALL = ["profit_dp","emission_dp","lexico_emissions_dp","lexico_profit_dp","profit_greedy"]

def cumulative_figure(year):
    print(f"\n  Building cumulative figure for {year} ...")
    dam = load_dam()

    # Load dispatch for all strategies
    data = {}
    for s in STRATS_ALL:
        p = os.path.join(PROC, f"annual_results_{year}_{s}.csv")
        if not os.path.exists(p):
            print(f"    WARNING: {p} not found - skipping {s}")
            continue
        df = pd.read_csv(p, index_col=0, parse_dates=True)
        p_yr = dam.reindex(df.index)
        cum_profit = ((df["discharge_mw"] * p_yr) - (df["charge_mw"] * p_yr)).cumsum()
        cum_co2    = df["net_em"].cumsum() / 1000.0  # tCO2
        data[s]    = (cum_profit, cum_co2)

    fig, (ax_top, ax_bot) = plt.subplots(2, 1, figsize=(12, 7), sharex=True)

    for s in STRATS_ALL:
        if s not in data: continue
        cp, ce = data[s]
        clr    = STRAT_COLOR[s]
        lbl    = STRAT_LABEL[s]
        ls     = STRAT_LS.get(s, "-")
        ax_top.plot(cp.index, cp.values, color=clr, label=lbl, lw=1.6, ls=ls)
        ax_bot.plot(ce.index, ce.values, color=clr, label=lbl, lw=1.6, ls=ls)

    # Month boundaries
    for m in pd.date_range(f"{year}-01-01", f"{year}-12-31", freq="MS"):
        ax_top.axvline(m, color="grey", lw=0.3, ls=":", alpha=0.5)
        ax_bot.axvline(m, color="grey", lw=0.3, ls=":", alpha=0.5)

    # Gas crisis band removed per paper style guide

    ax_top.axhline(0, color="black", lw=0.6)
    ax_bot.axhline(0, color="black", lw=0.6)
    ax_top.set_ylabel("Cumulative cashflow [EUR]")
    ax_bot.set_ylabel("Cumulative net CO₂ [tCO₂]")
    ax_bot.set_xlabel(str(year))

    if year == 2022:
        ax_top.set_title(f"Cumulative dispatch performance {year} – European gas crisis")
    else:
        ax_top.set_title(f"Cumulative dispatch performance {year}")

    ax_top.legend(fontsize=8, loc="upper left")
    ax_bot.legend(fontsize=8, loc="lower left")
    fig.tight_layout()
    savefig(fig, f"fig_cumulative_{year}")


def cumulative_figure_greedy(year):
    """Same layout as cumulative_figure(), isolated to the greedy benchmark.

    On the combined figure the greedy benchmark's cumulative curves are
    nearly flat compared to the DP strategies (it trades roughly 10x less
    volume), so they're hard to read against the shared axis scale. This
    figure re-plots just "profit_greedy" on its own axis scale.

    Reads model/run.py's own dispatch_profit_greedy.csv (single-year output,
    always current for whatever START_DATE/END_DATE run.py is configured
    for) rather than support/run_paper.py's annual_results_{year}_*.csv - 
    the latter is a separate, multi-year pipeline whose greedy outputs are
    only refreshed by a full support/run_paper.py rerun. Only usable for
    the year run.py is currently configured for (2022 by default).
    """
    print(f"\n  Building greedy-only cumulative figure for {year} ...")
    dam = load_dam()

    s = "profit_greedy"
    p = os.path.join(PROC, "dispatch_profit_greedy.csv")
    if not os.path.exists(p):
        print(f"    WARNING: {p} not found - skipping")
        return
    df = pd.read_csv(p, index_col=0, parse_dates=True)
    data_years = df.index.year.unique()
    if list(data_years) != [year]:
        print(f"    WARNING: dispatch_profit_greedy.csv covers {list(data_years)}, "
              f"not {year} - skipping (rerun model/run.py for {year} first)")
        return
    p_yr = dam.reindex(df.index)
    cum_profit = ((df["discharge_mw"] * p_yr) - (df["charge_mw"] * p_yr)).cumsum()
    cum_co2    = df["net_emissions_kg"].cumsum() / 1000.0  # tCO2

    fig, (ax_top, ax_bot) = plt.subplots(2, 1, figsize=(12, 7), sharex=True)

    clr = STRAT_COLOR[s]
    lbl = STRAT_LABEL[s]
    ax_top.plot(cum_profit.index, cum_profit.values, color=clr, label=lbl, lw=1.6)
    ax_bot.plot(cum_co2.index,    cum_co2.values,    color=clr, label=lbl, lw=1.6)

    for m in pd.date_range(f"{year}-01-01", f"{year}-12-31", freq="MS"):
        ax_top.axvline(m, color="grey", lw=0.3, ls=":", alpha=0.5)
        ax_bot.axvline(m, color="grey", lw=0.3, ls=":", alpha=0.5)

    ax_top.axhline(0, color="black", lw=0.6)
    ax_bot.axhline(0, color="black", lw=0.6)
    ax_top.set_ylabel("Cumulative cashflow [EUR]")
    ax_bot.set_ylabel("Cumulative net CO₂ [tCO₂]")
    ax_bot.set_xlabel(str(year))
    ax_top.set_title(f"Cumulative dispatch performance {year} – {lbl}")

    ax_top.legend(fontsize=8, loc="upper left")
    ax_bot.legend(fontsize=8, loc="lower left")
    fig.tight_layout()
    savefig(fig, f"fig_cumulative_greedy_{year}")


def greedy_mechanism_figure(year):
    """Greedy benchmark: cumulative profit + cumulative net CO2.

    Reads model/run.py's own dispatch_profit_greedy.csv + mei_curve.csv
    (always fresh for whichever year run.py is currently configured for),
    same data source as cumulative_figure_greedy(); see that function's
    docstring for why (avoids the stale support/run_paper.py per-year
    files unless those have just been regenerated).

    Does not modify cumulative_figure_greedy() or its output; this is a
    separate, additional figure.
    """
    print(f"\n  Building greedy mechanism figure for {year} ...")

    dispatch_path = os.path.join(PROC, "dispatch_profit_greedy.csv")
    mei_path      = os.path.join(PROC, "mei_curve.csv")
    if not (os.path.exists(dispatch_path) and os.path.exists(mei_path)):
        print(f"    WARNING: {dispatch_path} or {mei_path} not found, skipping")
        return

    df  = pd.read_csv(dispatch_path, index_col=0, parse_dates=True)
    mc  = pd.read_csv(mei_path, index_col=0, parse_dates=True)
    data_years = df.index.year.unique()
    if list(data_years) != [year]:
        print(f"    WARNING: dispatch_profit_greedy.csv covers {list(data_years)}, "
              f"not {year}, skipping (rerun model/run.py for {year} first)")
        return

    price = mc["dam_price_eur_mwh"].reindex(df.index)
    cum_profit = ((df["discharge_mw"] * price) - (df["charge_mw"] * price)).cumsum()
    # net_emissions_kg = charge_emissions_kg - discharge_avoided_kg: this is
    # already a net figure (emissions caused by charging, net of emissions
    # avoided by discharging), but it is operational only, it does not
    # subtract embodied/manufacturing emissions (see cem_generality.csv's
    # operational_net_co2_kg vs. full_accounting_net_co2_kg for that
    # distinction). Labelled "net operational" below to avoid ambiguity.
    cum_co2    = df["net_emissions_kg"].cumsum()   # kg, not tCO2, per spec

    fig, (ax_top, ax_bot) = plt.subplots(2, 1, figsize=(12, 7.5), sharex=True)
    clr = STRAT_COLOR["profit_greedy"]
    ax_top.plot(cum_profit.index, cum_profit.values, color=clr, lw=1.4, zorder=2,
                label="Greedy benchmark")
    ax_bot.plot(cum_co2.index, cum_co2.values, color=clr, lw=1.4, zorder=2,
                label="Greedy benchmark")

    for m in pd.date_range(f"{year}-01-01", f"{year}-12-31", freq="MS"):
        ax_top.axvline(m, color="grey", lw=0.3, ls=":", alpha=0.5)
        ax_bot.axvline(m, color="grey", lw=0.3, ls=":", alpha=0.5)
    ax_top.axhline(0, color="black", lw=0.6)
    ax_bot.axhline(0, color="black", lw=0.6)

    ax_top.set_ylabel("Cumulative cashflow [EUR]")
    ax_bot.set_ylabel("Cumulative net operational CO₂ [kg]")
    ax_bot.set_xlabel(str(year))
    ax_top.set_title(f"Greedy dispatch algorithm, {year}")
    ax_top.legend(fontsize=7, loc="upper left")

    fig.tight_layout()
    savefig(fig, f"fig_greedy_mechanism_{year}")


# Embodied-intensity generality sweep figure. Data produced by
# support/run_cem_generality_sweep.py -> cem_generality.csv; this
# function only reads that CSV and draws the figure.

def cem_generality_figure():
    """Figure - full-accounting CO2 reduction vs. embodied emission
    intensity c_em, years 2022 and 2024. Reads cem_generality.csv only
    (no hardcoded values, including the endpoint retention percentages,
    which come from that file's own retention_pct column).
    """
    print("\n  Building embodied-intensity generality figure ...")
    path = os.path.join(PROC, "cem_generality.csv")
    if not os.path.exists(path):
        print(f"    WARNING: {path} not found - skipping "
              f"(run support/run_cem_generality_sweep.py first)")
        return
    df = pd.read_csv(path)

    MM = 1 / 25.4
    fig, ax = plt.subplots(figsize=(174 * MM, 110 * MM), constrained_layout=True)

    colors = {2022: "#D62728", 2024: "#1F77B4"}
    curves = {}
    for year, g in df.groupby("year"):
        g = g.sort_values("c_em_kg_per_mwh")
        reduction = -g["full_accounting_net_co2_kg"] / 1000.0  # tCO2, positive = net benefit
        curves[year] = (g["c_em_kg_per_mwh"].values, reduction.values, g["retention_pct"].values)
        ax.plot(g["c_em_kg_per_mwh"], reduction, marker="o", markersize=5,
                lw=1.5, color=colors.get(year, "grey"), label=f"{year}", zorder=3)

    ax.axhline(0, color="black", lw=0.6, zorder=2)
    ax.axvline(20.3, color="grey", lw=0.8, ls="--", alpha=0.7, zorder=1)

    ax.set_xlabel("Embodied emission intensity  c_em  [kg CO₂eq / MWh discharged]")
    ax.set_ylabel("Full-accounting CO₂ reduction [tCO₂]")
    ax.grid(True, which="major", color="lightgrey", alpha=0.4, lw=0.5, zorder=0)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.legend(loc="upper right", fontsize=8, frameon=False)

    # Extra right-hand margin, explicit rather than relying on default
    # autoscale padding, so the endpoint retention-% labels (placed just
    # right of x=150, the last data point) have guaranteed room and can't
    # be clipped at the axes edge.
    xlo, xhi = ax.get_xlim()
    ax.set_xlim(xlo, xhi + 0.12 * (xhi - xlo))

    # "paper default" label: vertically centred in the plot area, just right
    # of the reference line, positioned after xlim/ylim are finalised so
    # it's actually centred (not anchored to a data value that could clip).
    xlo, xhi = ax.get_xlim()
    ylo, yhi = ax.get_ylim()
    label_x = 20.3 + 0.015 * (xhi - xlo)
    ax.text(label_x, (ylo + yhi) / 2, "paper default\n(20.3 kg/MWh)",
            fontsize=7, color="grey", ha="left", va="center", linespacing=1.4)

    # Endpoint retention-% labels at x=150 (the last c_em value in the
    # sweep), read directly from retention_pct - not hardcoded. Offset in
    # points (not data units) so the two labels clear each other regardless
    # of how close the underlying y-values are.
    offsets = {2022: -8, 2024: 8}
    for year, (xs, ys, pct) in curves.items():
        x_end, y_end, pct_end = xs[-1], ys[-1], pct[-1]
        ax.annotate(f"{pct_end:.0f}%", xy=(x_end, y_end),
                    xytext=(6, offsets.get(year, 0)), textcoords="offset points",
                    fontsize=7, color=colors.get(year, "grey"),
                    ha="left", va="center")

    savefig(fig, "fig_cem_generality")

    # ---- VERIFY: print the values used for the endpoint annotations ----
    print("  Verification - y-values and retention % at c_em=150:")
    for year, (xs, ys, pct) in curves.items():
        print(f"    {year}: y={ys[-1]:.2f} tCO2, retention={pct[-1]:.2f}% "
              f"(annotated as \"{pct[-1]:.0f}%\")")


# Step 5: CO2 price over model horizon

def co2_price_horizon_figure():
    print("\n  Building CO2 price horizon figure ...")
    prices = load_prices()
    co2 = prices["CO2"].dropna()

    if co2.empty:
        print("  WARNING: CO2 series empty - skipping fig_co2_price_horizon")
        return

    fig, ax = plt.subplots(figsize=(12, 4))
    ax.plot(co2.index, co2.values, color="#C62828", lw=1.2, label="CO2 EUA")

    for y in pd.date_range(co2.index.min().normalize(), co2.index.max().normalize(), freq="YS"):
        ax.axvline(y, color="black", lw=0.3, ls=":", alpha=0.35)

    ax.set_ylabel("CO2 price [EUR/t]")
    ax.set_xlabel("Date")
    ax.set_title(
        f"CO2 price over model horizon; "
        f"{co2.index.min().strftime('%d-%m-%Y')} to {co2.index.max().strftime('%d-%m-%Y')}"
    )
    ax.legend(loc="upper left", fontsize=8, framealpha=0.9)
    fig.tight_layout()
    savefig(fig, "fig_co2_price_horizon")


# Step 6: annual KPI table

def annual_kpi_table():
    print("\n  Building annual KPI table ...")
    base = pd.read_csv(os.path.join(PROC, "table_annual_kpi_summary.csv"))

    strats = ["profit_dp","emission_dp","lexico_emissions_dp","lexico_profit_dp","profit_greedy"]
    rows   = []
    for _, r in base.iterrows():
        row = {"Year": int(r["year"])}
        for s in strats:
            p  = r.get(f"{s}_profit_eur",    r.get(f"{s}_total_profit_eur", 0))
            c  = r.get(f"{s}_net_co2_kg",     r.get(f"{s}_net_emissions_kg_co2", 0))
            flag = "*" if c > 0 else ""
            row[f"{STRAT_LABEL[s]} profit (EUR)"]  = f"{int(round(p)):,}"
            # The flag must go in the cell value, not the column name -
            # putting it in the column name would split this into two
            # columns (with vs. without "*") since `flag` varies row-to-row.
            row[f"{STRAT_LABEL[s]} net CO2 (kg)"] = f"{int(round(c)):,}{flag}"
        rows.append(row)

    df_out = pd.DataFrame(rows)
    out_path = os.path.join(PROC, "table_annual_kpi_paper.csv")
    df_out.to_csv(out_path, index=False)

    # Console print
    print()
    print("  Annual KPI Summary:")
    print(f"  {'Year':<6}", end="")
    for s in strats:
        lbl = STRAT_LABEL[s][:12]
        print(f"  {lbl+' EUR':>16}  {lbl+' CO2':>14}", end="")
    print()
    print("  " + "-" * 130)
    for _, r in base.iterrows():
        print(f"  {int(r['year']):<6}", end="")
        for s in strats:
            p = r.get(f"{s}_profit_eur", r.get(f"{s}_total_profit_eur", 0))
            c = r.get(f"{s}_net_co2_kg", r.get(f"{s}_net_emissions_kg_co2", 0))
            flag = "*" if c > 0 else " "
            print(f"  {int(round(p)):>16,}  {int(round(c)):>13,}{flag}", end="")
        print()
    print()
    print("  * = net CO2 added  (positive = storage increased grid emissions)")
    saved_figures.append(out_path)
    return out_path


# Step 8: formatted tables (.xlsx, for manuscript authoring)
# Ported from make_figures.py (merged into this file - see module docstring).

def xl_header(ws, row, cols, col_widths):
    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF", size=9)
    thin  = Side(style="thin", color="BFBFBF")
    bord  = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, (h, w) in enumerate(zip(cols, col_widths), 1):
        cell = ws.cell(row, ci, h)
        cell.font = hfont; cell.fill = hfill
        cell.border = bord; cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[row].height = 22


def xl_row(ws, row_idx, vals, col_widths, fill_color=None):
    thin = Side(style="thin", color="BFBFBF")
    bord = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill("solid", fgColor=fill_color) if fill_color else None
    for ci, v in enumerate(vals, 1):
        cell = ws.cell(row_idx, ci, v)
        cell.font = Font(size=9)
        cell.border = bord
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if fill: cell.fill = fill
    ws.row_dimensions[row_idx].height = 20


# Table 1: Schram comparison
def make_table1():
    print("  Table 1: Schram comparison ...")
    rows_data = [
        ("Eemshaven A/B (RWE)","Coal",27.5,32.9,743,769,"Reference fuel price (2014 vs 2018 actual); Emission factor source"),
        ("MPP3 / Maasvlakte","Coal",27.0,32.9,728,769,"Reference fuel price (2014 vs 2018 actual); Emission factor source"),
        ("Amer (Amercentrale)","Coal/Biomass",31.2,31.8,852,406,"Biomass co-firing 2018 (50% biomass, zero CO2)"),
        ("Sloecentrale","Gas CCGT",44.1,36.1,347,337,"Reference fuel price (2014 vs 2018 actual)"),
        ("Magnum Eemshaven","Gas CCGT",44.6,37.3,351,349,"Reference fuel price (2014 vs 2018 actual)"),
        ("Claus C / Maasbracht","Gas CCGT",46.2,37.3,364,349,"Reference fuel price (2014 vs 2018 actual)"),
        ("Velsen 24","BFG",74.2,6.9,728,634,"Efficiency value; fuel cost treatment (waste gas = 0 EUR/MWh_th)"),
    ]
    cols    = ["Plant","Fuel","Schram SRMC\n(EUR/MWh)","Model SRMC\n(EUR/MWh)","Schram Em.\n(kg CO2/MWh)","Model Em.\n(kg CO2/MWh)","Primary reason for difference"]
    col_w   = [24, 12, 14, 14, 14, 14, 42]
    FUEL_FILL = {"Coal":"F2DCDB","Coal/Biomass":"EBF1DE","Gas CCGT":"DAEEF3","BFG":"FFF2CC"}

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Table 1"
    xl_header(ws, 1, cols, col_w)
    for ri, r in enumerate(rows_data, 2):
        fc = FUEL_FILL.get(r[1])
        xl_row(ws, ri, list(r), col_w, fc)
    out = os.path.join(PROC, "table1_schram_comparison.xlsx")
    wb.save(out); saved_figures.append(out); print(f"    Saved: {out}")


# Table 2: Plant overview (rerunnable from CSV)
def make_table2():
    print("  Table 2: Plant overview ...")
    rows_data = [
        ("Eemshaven A/B (RWE)","Coal (supercritical)",1560,0.460,"Supercritical coal (>=46%)","[1]",""),
        ("Maasvlakte 3 (MPP3)","Coal (supercritical)",1100,0.460,"Supercritical coal (>=46%)","[6, 7]",""),
        ("Rotterdam 1 (Onyx)","Coal (supercritical)",736,0.450,"Supercritical coal (>=46%)","[8]",""),
        ("Amer (Amercentrale)","Biomass (co-firing)",540,0.436,"Subcritical biomass/coal","[35]","Must-run note"),
        ("Flevo 4/5 (Maxima)","Gas CCGT",940,0.600,"H-class CCGT (>=60%)","[13]",""),
        ("Diemen 34 CCGT","Gas CCGT",435,0.590,"F-class CCGT (56-60%)","[15]",""),
        ("Enecogen 10/20","Gas CCGT",950,0.590,"F-class CCGT (56-60%)","[10]",""),
        ("Hemweg 9","Gas CCGT",440,0.590,"F-class CCGT (56-60%)","[11, 15]",""),
        ("Sloecentrale 10/20","Gas CCGT",940,0.600,"F-class/upgraded CCGT (>=60%)","[9]",""),
        ("Claus C 2/3/4/5","Gas CCGT",1320,0.580,"H-class CCGT (>=60%)","[12, 3]",""),
        ("Eemshaven 10/20/30 (Magnum)","Gas CCGT",1455,0.580,"F-class CCGT (56-60%)","[2]",""),
        ("Moerdijk 2 CCGT","Gas CCGT",426,0.580,"F-class CCGT (56-60%)","[4, 16]",""),
        ("Maasstroom Rijnmond 2","Gas CCGT",426,0.520,"F-class CCGT (single-shaft)","[14]",""),
        ("Eems CCGT 4/5/6/7","Gas CCGT",1800,0.560,"F-class CCGT (56-60%)","[32]",""),
        ("Elsta Cogeneration","Gas CHP",459,0.547,"CHP (electrical only)","[23]","Must-run note"),
        ("Moerdijk 1 CHP","Gas CHP",348,0.580,"CHP (electrical only)","[30, 4]","Must-run note"),
        ("Rijnmond 1","Gas CHP",810,0.580,"CHP (electrical only)","[22]","Must-run note"),
        ("Diemen 33 CHP","Gas CCGT",249,0.502,"CHP/CCGT peaker","[21]","Must-run note"),
        ("Delesto 2","Gas CHP",523,0.550,"CHP (electrical only)","[28]","Must-run note"),
        ("Lage Weide 6","Gas CHP",248,0.547,"CHP (electrical only)","[25]","Must-run note"),
        ("Merwedekanaal Pegus 12","Gas CHP",233,0.502,"CHP (electrical only)","[24]","Must-run note"),
        ("NAM Schoonebeek","Gas CHP",128,0.502,"CHP (electrical only)","[33]","Must-run note"),
        ("Den Haag Power station","Gas CHP",112,0.440,"CHP (aero-derivative)","[26, 27]","Must-run note"),
        ("RoCa","Gas CHP",264,0.482,"CHP (electrical only)","[29]","Must-run note"),
        ("Pergen 1/2","Gas CHP",308,0.482,"CHP (electrical only)","[21]","Must-run note"),
        ("Swentibold Chemelot","Gas CHP",230,0.480,"CHP (electrical only)","[5]","Must-run note"),
        ("Velsen 25","Blast furnace gas",380,0.448,"BFG (Schram methodology)","[17, 18, 19]","Must-run note"),
        ("Velsen 24","Blast furnace gas",460,0.399,"BFG (Schram methodology)","[17, 18, 19]","Must-run note"),
        ("Bergum 10GT/20GT Friesland","Gas OCGT",152,0.350,"Modern OCGT","[20]",""),
        ("Eems 20","Gas OCGT",131,0.344,"Older OCGT (<40%)","[21]",""),
        ("Borssele 30","Nuclear",486,0.355,"Nuclear (thermal eff ~35%)","[34]",""),
        ("RES","Solar + Wind",27000,1.000,"Renewable (zero marginal cost)","[31]",""),
        ("Hemweg 8","Coal (subcritical)",630,0.420,"Subcritical coal (~42%)","[36, 37, 38]","Decommissioned 20-12-2019"),
    ]
    cols  = ["Plant","Fuel / Technology","Cap. (MW)","η (net)","Category","Ref.","Notes"]
    col_w = [34, 22, 10, 8, 30, 12, 28]
    FUEL_FILL = {
        "Coal":"F2DCDB","Coal (supercritical)":"F2DCDB","Coal (subcritical)":"F2DCDB",
        "Biomass":"EBF1DE","Biomass (co-firing)":"EBF1DE",
        "Gas CCGT":"DAEEF3","Gas CHP":"E2EFDA","Gas OCGT":"EDEDED",
        "Blast furnace gas":"FFF2CC","Nuclear":"E4DFEC","Solar + Wind":"D9F0D3",
    }
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Table 2"
    xl_header(ws, 1, cols, col_w)
    for ri, r in enumerate(rows_data, 2):
        fc = FUEL_FILL.get(r[1])
        xl_row(ws, ri, list(r), col_w, fc)
    out = os.path.join(PROC, "table2_plant_overview.xlsx")
    wb.save(out); saved_figures.append(out); print(f"    Saved: {out}")
    print("    NOTE: To update citations, edit the rows_data list in make_table2() and rerun.")


# Table 3: Battery parameters
def make_table3():
    print("  Table 3: Battery parameters ...")
    rows_data = [
        ("Rated power","P_max","1.016","MW","Huawei (2024)"),
        ("Energy capacity","E_cap","2.032","MWh","Huawei (2024)"),
        ("Round-trip efficiency","eta_RT","0.913","-","Huawei (2024)"),
        ("Charge efficiency","eta_c","0.956","-","sqrt(eta_RT)"),
        ("Discharge efficiency","eta_d","0.956","-","sqrt(eta_RT)"),
        ("Minimum SOC","SOC_min","0.0","MWh","-"),
        ("Maximum SOC","SOC_max","2.032","MWh","-"),
        ("Initial SOC","SOC_0","0.0","MWh","-"),
        ("Cycle cost proxy","c_cycle","55.75","EUR/MWh","Huawei (2024); IRENA (2024)"),
    ]
    cols  = ["Parameter","Symbol","Value","Unit","Source"]
    col_w = [28, 14, 10, 10, 32]
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Table 3"
    xl_header(ws, 1, cols, col_w)
    alt = ["F2F2F2", None]
    for ri, r in enumerate(rows_data, 2):
        fc = alt[(ri-2)%2]
        xl_row(ws, ri, list(r), col_w, fc)
    out = os.path.join(PROC, "table3_battery_parameters.xlsx")
    wb.save(out); saved_figures.append(out); print(f"    Saved: {out}")


# Table 4: Annual KPI (formatted .xlsx companion to annual_kpi_table()'s CSV)
def make_table4():
    print("  Table 4: Annual KPI summary (xlsx) ...")
    base = pd.read_csv(os.path.join(PROC, "table_annual_kpi_summary.csv"))
    strats = ["profit_dp","emission_dp","lexico_emissions_dp","lexico_profit_dp","profit_greedy"]
    labels = ["Profit max (DP)","Emission min (DP)","Lexico-E","Lexico-P","Greedy"]

    cols = ["Year"]
    for lbl in labels:
        cols += [f"{lbl}\nProfit (EUR)", f"{lbl}\nNet CO2 (kg)"]
    col_w = [7] + [16, 14]*len(strats)

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Table 4"
    xl_header(ws, 1, cols, col_w)
    ws.row_dimensions[1].height = 30

    CO2ADD = PatternFill("solid", fgColor="FFCCCC")
    ri = 1
    for ri, (_, r) in enumerate(base.iterrows(), 2):
        yr = int(r["year"])
        vals = [yr]
        for s in strats:
            p = r.get(f"{s}_profit_eur", r.get(f"{s}_total_profit_eur", 0))
            c = r.get(f"{s}_net_co2_kg", r.get(f"{s}_net_emissions_kg_co2", 0))
            flag = " *" if c > 0 else ""
            vals += [f"{int(round(p)):,}", f"{int(round(c)):,}{flag}"]
        xl_row(ws, ri, vals, col_w)
        # Highlight positive CO2 cells
        for ci in range(3, len(cols)+1, 2):
            v = str(ws.cell(ri, ci).value or "")
            if "*" in v:
                ws.cell(ri, ci).fill = CO2ADD

    # Add legend row
    ws.cell(ri+2, 1, "* = net CO2 added (storage increased grid emissions for this year)").font = Font(size=8, italic=True)

    out = os.path.join(PROC, "table4_annual_kpi.xlsx")
    wb.save(out); saved_figures.append(out); print(f"    Saved: {out}")


# Figure 2: workflow diagram (schematic - matplotlib only, no graphviz).
# Monochrome, minimalist, 5-box left-to-right flow: white background,
# black outlines only, no fill colour, no shadows or icons.

_WF_INK = "#000000"


def _draw_box(ax, x, y, w, h, lw=0.9, radius=1.6, zorder=2):
    """A box in the pure black-outline-on-white style: no fill, thin edge."""
    box = mpatches.FancyBboxPatch(
        (x, y), w, h,
        boxstyle=f"round,pad=0,rounding_size={radius}",
        facecolor="white", edgecolor=_WF_INK, linewidth=lw, zorder=zorder,
    )
    ax.add_patch(box)
    return box


def _draw_arrow(ax, xy_from, xy_to, lw=1.0, zorder=3):
    """A thin, straight, black arrow between two points."""
    arr = mpatches.FancyArrowPatch(
        xy_from, xy_to, arrowstyle="-|>", mutation_scale=7,
        linewidth=lw, color=_WF_INK, zorder=zorder,
        shrinkA=0, shrinkB=0,
    )
    ax.add_patch(arr)
    return arr


def framework_figure():
    """Figure 2 - model workflow, Model Input -> Merit-order Inversion ->
    Marginal Emissions Intensity Curve -> Storage Dispatch -> Model Outputs,
    drawn as five identically-sized boxes in a straight left-to-right chain.

    Monochrome by design (no colour channel at all, so none of the
    categorical-palette validation elsewhere in this file applies here) - 
    black FancyBboxPatch outlines on white, thin FancyArrowPatch connectors,
    bold sans-serif titles, plain bullet lists. Every box is the same
    (w, h); every wrapped bullet line below was sized against actual
    rendered text-extent measurements at 8pt (not estimated) so five
    same-size boxes can hold visibly different amounts of content without
    any of them overflowing.

    Output: fig2_framework.pdf (vector) + .png (300dpi preview), via the
    shared savefig() helper.
    """
    print("\n  Building workflow diagram figure (Figure 2) ...")

    MM = 1 / 25.4
    BOX_W, BOX_H = 46.0, 55.0
    GAP = 14.0
    MARGIN = 3.0
    N = 5
    W = 2 * MARGIN + N * BOX_W + (N - 1) * GAP
    H = 2 * MARGIN + BOX_H

    fig = plt.figure(figsize=(W * MM, H * MM))
    ax = fig.add_axes([0, 0, 1, 1])
    ax.set_xlim(0, W)
    ax.set_ylim(0, H)
    ax.axis("off")

    FS_TITLE, FS_BULLET = 9, 8
    box_y = MARGIN
    title_reserve = 14.0   # fixed for all 5 boxes -> bullets align across panels
    bullet_line_h = 3.6
    bullet_gap = 3.0
    pad_x = 3.0

    columns = [
        dict(
            title="Model Input",
            bullets=[
                ["Continuous electricity", "spot prices"],
                ["Continuous commodity", "spot prices"],
            ],
        ),
        dict(
            title="Merit-order Inversion",
            bullets=[
                ["Power-plant efficiency", "data"],
                ["Commodity prices"],
                ["Continuous merit-order", "stack"],
            ],
        ),
        dict(
            title="Marginal Emissions\nIntensity Curve",
            bullets=[
                ["Power-plant efficiency", "data"],
                ["Electricity spot prices"],
                ["Continuous marginal", "emissions intensity", "signal"],
            ],
        ),
        dict(
            title="Storage Dispatch",
            bullets=[
                ["Technology parameters"],
                ["User-defined dispatch", "algorithm"],
                ["Storage representation", "within the system"],
            ],
        ),
        dict(
            title="Model Outputs",
            bullets=[
                ["Realized profits/costs"],
                ["Realized CO₂ emissions"],
            ],
        ),
    ]

    box_lefts = [MARGIN + i * (BOX_W + GAP) for i in range(N)]

    for i, col in enumerate(columns):
        x = box_lefts[i]
        _draw_box(ax, x, box_y, BOX_W, BOX_H)

        ax.text(x + BOX_W / 2, box_y + BOX_H - 3, col["title"],
                ha="center", va="top", fontsize=FS_TITLE, fontweight="bold",
                color=_WF_INK, linespacing=1.3, fontfamily="sans-serif")

        y_cursor = box_y + BOX_H - title_reserve
        for bullet_lines in col["bullets"]:
            label = "•  " + bullet_lines[0]
            if len(bullet_lines) > 1:
                label += "\n" + "\n".join("   " + ln for ln in bullet_lines[1:])
            ax.text(x + pad_x, y_cursor, label, ha="left", va="top",
                    fontsize=FS_BULLET, color=_WF_INK, linespacing=1.25,
                    fontfamily="sans-serif")
            y_cursor -= len(bullet_lines) * bullet_line_h + bullet_gap

        if i < N - 1:
            y_mid = box_y + BOX_H / 2
            _draw_arrow(ax, (x + BOX_W, y_mid), (box_lefts[i + 1], y_mid))

    savefig(fig, "fig2_framework")
    print(f"    Saved: fig2_framework.pdf / .png")


# Main

if __name__ == "__main__":
    print("=" * 65)
    print("PRODUCING ALL PAPER FIGURES")
    print("=" * 65)

    # Step 0
    print("\nSTEP 0: Workflow diagram figure (Figure 2)")
    framework_figure()

    # Step 1
    print("\nSTEP 1: Merit order stacked bar figures")
    co2_18, gas_18, coal_18 = merit_order_figure(
        "2018-01-11",
        subtitle=None,
        fname="fig_merit_order_20180111"
    )
    co2_22, gas_22, coal_22 = merit_order_figure(
        "2022-08-15",
        subtitle=None,
        fname="fig_merit_order_20220815"
    )
    co2_22w, gas_22w, coal_22w = merit_order_figure(
        "2022-08-26",
        subtitle=None,
        fname="fig_merit_order_20220826"
    )
    co2_25w, gas_25w, coal_25w = merit_order_figure(
        "2025-08-26",
        subtitle=None,
        fname="fig_merit_order_20250826"
    )

    # Step 1b
    print("\nSTEP 1b: Schram et al. (2019) validation (Figure 4)")
    schram_validation_figure()

    # Step 2
    print("\nSTEP 2: MEI time series (dispatch week)")
    mei_dispatch_week_figure()

    # Step 3
    print("\nSTEP 3: Dispatch detail figure")
    existing_dd = os.path.join(FIG_DIR, "fig_dispatch_detail_week.png")
    if os.path.exists(existing_dd):
        print(f"  Already exists: {existing_dd} - regenerating with current style")
    dispatch_detail_figure()

    # Step 4
    print("\nSTEP 4: Cumulative figures (2022 and 2025)")
    for yr in [2022, 2025]:
        existing = os.path.join(FIG_DIR, f"fig_cumulative_{yr}.png")
        if os.path.exists(existing):
            print(f"  fig_cumulative_{yr}.png already exists - regenerating")
        cumulative_figure(yr)
        cumulative_figure_greedy(yr)

    print("\nSTEP 4b: Greedy mechanism figure (2022)")
    greedy_mechanism_figure(2022)

    print("\nSTEP 4c: Embodied-intensity generality figure")
    cem_generality_figure()

    # Step 5
    print("\nSTEP 5: CO2 price horizon")
    co2_price_horizon_figure()

    # Step 6
    print("\nSTEP 6: Annual KPI table")
    annual_kpi_table()

    # Step 8: Formatted tables (.xlsx)
    print("\nSTEP 8: Formatted tables (xlsx)")
    make_table1()
    make_table2()
    make_table3()
    make_table4()

    # Step 7: Console report
    print("\n" + "=" * 65)
    print("STEP 7: CONSOLE REPORT")
    print("=" * 65)
    print(f"\nFigures successfully produced ({len(saved_figures)}):")
    for f in saved_figures:
        print(f"  {f}")
    if skipped_figures:
        print(f"\nFigures skipped (already existed):")
        for f in skipped_figures: print(f"  {f}")
    if failed_figures:
        print(f"\nFigures that could not be produced:")
        for f, r in failed_figures: print(f"  {f}  ({r})")

    print(f"\nDAM prices at hour 14:00:")
    dam_all = load_dam()
    p18 = dam_all.get(pd.Timestamp("2018-01-11 14:00"), float("nan"))
    p22 = dam_all.get(pd.Timestamp("2022-08-15 14:00"), float("nan"))
    p22w = dam_all.get(pd.Timestamp("2022-08-26 14:00"), float("nan"))
    p25w = dam_all.get(pd.Timestamp("2025-08-26 14:00"), float("nan"))
    print(f"  2018-01-11 14:00: {p18:.2f} EUR/MWh")
    print(f"  2022-08-15 14:00: {p22:.2f} EUR/MWh")
    print(f"  2022-08-26 14:00: {p22w:.2f} EUR/MWh")
    print(f"  2025-08-26 14:00: {p25w:.2f} EUR/MWh")

    print(f"\nCommodity prices used for merit order figures:")
    print(f"  2018-01-11: CO2={co2_18:.2f} EUR/t | Gas={gas_18:.2f} | Coal={coal_18:.2f} EUR/MWh_th")
    print(f"  2022-08-15: CO2={co2_22:.2f} EUR/t | Gas={gas_22:.2f} | Coal={coal_22:.2f} EUR/MWh_th")
    print(f"  2022-08-26: CO2={co2_22w:.2f} EUR/t | Gas={gas_22w:.2f} | Coal={coal_22w:.2f} EUR/MWh_th")
    print(f"  2025-08-26: CO2={co2_25w:.2f} EUR/t | Gas={gas_25w:.2f} | Coal={coal_25w:.2f} EUR/MWh_th")

    print(f"\nDispatch week: {DISPATCH_WEEK_START} to {DISPATCH_WEEK_END}")

    print(f"\nHourly dispatch data availability:")
    for yr in range(2018, 2026):
        p = os.path.join(PROC, f"annual_results_{yr}_profit_dp.csv")
        print(f"  {yr}: {'EXISTS' if os.path.exists(p) else 'MISSING'}")

    print("\nDone.")
